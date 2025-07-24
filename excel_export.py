from openpyxl import Workbook, load_workbook
import os

def export_to_excel(filename, survey_record):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        headers = [
            "User Name", "Address", "Mail ID", "Mobile No", "Date", "Survey Title", "Survey Category", "Survey Use Case"
        ] + [f"Q{i}" for i in range(1, 11)] \
          + [f"Answer{i}" for i in range(1, 11)] \
          + [f"AI Review{i}" for i in range(1, 11)] \
          + [f"AI Recommendation{i}" for i in range(1, 11)]
        ws.append(headers)
    user = survey_record['user_info']
    base_data = [
        user.get('Name',''), user.get('Address',''), user.get('Email',''), user.get('Phone',''), user.get('Date',''),
        survey_record.get('survey_title',''),
        survey_record.get('survey_category',''),
        survey_record.get('survey_use_case','')
    ]
    questions      = survey_record.get('survey_questions', [])
    answers        = survey_record.get('user_answers', [])
    ai_review_list = survey_record.get('ai_review_list', [])
    q_data = questions + [""]*(10-len(questions))
    a_data = answers   + [""]*(10-len(answers))
    review_data  = [d.get("review_answer","") for d in ai_review_list] + [""]*(10-len(ai_review_list))
    recom_data   = [d.get("recommend_answer","") for d in ai_review_list]+ [""]*(10-len(ai_review_list))
    ws.append(base_data + q_data + a_data + review_data + recom_data)
    wb.save(filename)
